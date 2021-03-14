from openpyxl import load_workbook
import win32ui
def read_excel():
    global filename
    wrong = 0
    print("请打开要检验的噩梦版Excel文件")
    dlg = win32ui.CreateFileDialog(1, None, None, 0, "Excel files|*.xls*")  # 表示打开文件对话框
    dlg.SetOFNInitialDir('D:/')  # 设置打开文件对话框中的初始显示目录
    dlg.DoModal()
    filename = dlg.GetPathName()  # 获取选择的文件名称

def make_1008():
    workbook = load_workbook(filename,data_only=True)
    if "1008" in filename:
        sheet1 = workbook["1008"]
    GEN_HTML = "1008.html"
    f = open(GEN_HTML, 'w', encoding='utf-8')
    message = """
    <html xmlns:v="urn:schemas-microsoft-com:vml"
    xmlns:o="urn:schemas-microsoft-com:office:office"
    xmlns:x="urn:schemas-microsoft-com:office:excel"
    xmlns="http://www.w3.org/TR/REC-html40">

    <head>
        <script type="text/javascript">
//在页面未加载完毕之前显示的loading Html自定义内容
var _LoadingHtml = `<div id="loadingDiv">页面加载中，请等待...</div>`;
//呈现loading效果
document.write(_LoadingHtml);
//监听加载状态改变
document.body.style.display = "none";
document.onreadystatechange = completeLoading;
   
//加载状态为complete时移除loading效果
function completeLoading() {
    if (document.readyState == "complete") {
        document.body.style.display = "block";
        var loadingMask = document.getElementById('loadingDiv');
        loadingMask.parentNode.removeChild(loadingMask);
    } 
   else
       document.body.style.display = "none";
}
</script>

    <title>【噩梦版】WYF's 全缓冲角1008</title>
    <meta http-equiv=Content-Type content="text/html; charset=gb2312">
    <meta name=ProgId content=Excel.Sheet>
    <meta name=Generator content="Microsoft Excel 15">
    <link rel=File-List href="1008.files/filelist.xml">
    <!--[if !mso]>
    <style>
    v\:* {behavior:url(#default#VML);}
    o\:* {behavior:url(#default#VML);}
    x\:* {behavior:url(#default#VML);}
    .shape {behavior:url(#default#VML);}
    </style>
    <![endif]-->
    <style id="1008_Styles">
    <!--table
    	{mso-displayed-decimal-separator:"\.";
    	mso-displayed-thousand-separator:"\,";}
    .font1
    	{color:windowtext;
    	font-size:9.0pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:宋体;
    	mso-generic-font-family:auto;
    	mso-font-charset:134;}
    .font2
    	{color:black;
    	font-size:13.3pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:Tahoma, sans-serif;
    	mso-font-charset:0;}
    .xl1
    	{padding:0px;
    	mso-ignore:padding;
    	color:windowtext;
    	font-size:13.3pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:宋体;
    	mso-generic-font-family:auto;
    	mso-font-charset:134;
    	mso-number-format:General;
    	text-align:general;
    	vertical-align:middle;
    	background:white;
    	mso-pattern:black none;
    	mso-protection:unlocked visible;
    	white-space:nowrap;}
    .xl2
    	{padding:0px;
    	mso-ignore:padding;
    	color:windowtext;
    	font-size:13.3pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:宋体;
    	mso-generic-font-family:auto;
    	mso-font-charset:134;
    	mso-number-format:General;
    	text-align:general;
    	vertical-align:middle;
    	mso-background-source:auto;
    	mso-pattern:auto;
    	mso-protection:unlocked visible;
    	white-space:nowrap;}
    .xl3
    	{padding:0px;
    	mso-ignore:padding;
    	color:black;
    	font-size:13.3pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:Tahoma, sans-serif;
    	mso-font-charset:0;
    	mso-number-format:General;
    	text-align:center;
    	vertical-align:middle;
    	border:.5pt solid windowtext;
    	background:white;
    	mso-pattern:black none;
    	white-space:nowrap;}
    .xl4
    	{padding:0px;
    	mso-ignore:padding;
    	color:black;
    	font-size:13.3pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:Tahoma, sans-serif;
    	mso-font-charset:0;
    	mso-number-format:General;
    	text-align:center;
    	vertical-align:middle;
    	border:.5pt solid windowtext;
    	background:#D9D9D9;
    	mso-pattern:black none;
    	white-space:nowrap;}
    ruby
    	{ruby-align:left;}
    rt
    	{color:windowtext;
    	font-size:9.0pt;
    	font-weight:400;
    	font-style:normal;
    	text-decoration:none;
    	font-family:宋体;
    	mso-generic-font-family:auto;
    	mso-font-charset:134;
    	mso-char-type:none;}
    -->
    </style>
    </head>

    <body>
    <!--[if !excel]>　　<![endif]-->
    <!--下列信息由 Microsoft Excel 的发布为网页向导生成。-->
    <!--如果同一条目从 Excel 中重新发布，则所有位于 DIV 标记之间的信息均将被替换。-->
    <!----------------------------->
    <!--“从 EXCEL 发布网页”向导开始-->
    <!----------------------------->

    <div id="1008" align=center x:publishsource="Excel">
    
    <table border=0 cellpadding=0 cellspacing=0 width=1293 class=xl2
     style='border-collapse:collapse;table-layout:fixed;width:923.4pt'>
     <col class=xl1 width=53 style='mso-width-source:userset;mso-width-alt:
     1685;width:38pt'>
     <col class=xl1 width=378 style='mso-width-source:userset;mso-width-alt:
     12096;width:269.8pt'>
     <col class=xl1 width=53 style='mso-width-source:userset;mso-width-alt:
     1685;width:38pt'>
     <col class=xl1 width=378 style='mso-width-source:userset;mso-width-alt:
     12096;width:269.8pt'>
     <col class=xl1 width=53 style='mso-width-source:userset;mso-width-alt:
     1685;width:38pt'>
     <col class=xl1 width=378 style='mso-width-source:userset;mso-width-alt:
     12096;width:269.8pt'>
     <tr height=48 style='mso-height-source:userset;height:36.0pt'>
      <td height=48 class=xl3 width=53 style='height:36.0pt;width:40pt'><output id="""+'"'+sheet1.cell(1, 1).value+'"'+""" class=xl3 width=53 style='border-left:none;border-right:none;border-top:none;border-bottom:none;width:40pt'></td>
      <td class=xl3 width=378 style='border-left:none;width:284pt' title="""+'"'+sheet1.cell(1, 2).comment.text+'"'+""">"""+str(sheet1.cell(1, 2).value)+"""</td>
      <td class=xl3 width=53 style='border-left:none;width:40pt'><output id="""+'"'+sheet1.cell(1, 3).value+'"'+""" class=xl3 width=53 style='border-left:none;border-right:none;border-top:none;border-bottom:none;width:40pt'></td>
      <td class=xl3 width=378 style='border-left:none;width:284pt' title="""+'"'+sheet1.cell(1, 4).comment.text+'"'+""">"""+str(sheet1.cell(1, 4).value)+"""</td>
      <td class=xl3 width=53 style='border-left:none;width:40pt'><output id="""+'"'+sheet1.cell(1, 5).value+'"'+""" class=xl3 width=53 style='border-left:none;border-right:none;border-top:none;border-bottom:none;width:40pt'></td>
      <td class=xl3 width=378 style='border-left:none;width:284pt' title="""+'"'+sheet1.cell(1, 6).comment.text+'"'+""">"""+str(sheet1.cell(1, 6).value)+"""</td>
     </tr>
    """
    for i in range(2,127):
        message=message+""" 
         <tr height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl3 style='height:36.0pt;border-top:none'><output id="""+'"'+sheet1.cell(i, 1).value+'"'+""" height=48 class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 2).comment.text+'"'+""">"""+str(sheet1.cell(i, 2).value)+"""</td>
          <td class=xl3 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 3).value+'"'+""" class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 4).comment.text+'"'+""">"""+str(sheet1.cell(i, 4).value)+"""</td>
          <td class=xl3 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 5).value+'"'+""" class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 6).comment.text+'"'+""">"""+str(sheet1.cell(i, 6).value)+"""</td>
         </tr>
        """
    for i in range(127,217):
        message=message+""" 
         <tr height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl4 style='height:36.0pt;border-top:none'><output id="""+'"'+sheet1.cell(i, 1).value+'"'+""" height=48 class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 2).comment.text+'"'+""">"""+str(sheet1.cell(i, 2).value)+"""</td>
          <td class=xl4 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 3).value+'"'+""" class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 4).comment.text+'"'+""">"""+str(sheet1.cell(i, 4).value)+"""</td>
          <td class=xl4 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 5).value+'"'+""" class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 6).comment.text+'"'+""">"""+str(sheet1.cell(i, 6).value)+"""</td>
         </tr>
        """
    for i in range(217,277):
        message=message+""" 
         <tr height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl3 style='height:36.0pt;border-top:none'><output id="""+'"'+sheet1.cell(i, 1).value+'"'+""" height=48 class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 2).comment.text+'"'+""">"""+str(sheet1.cell(i, 2).value)+"""</td>
          <td class=xl3 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 3).value+'"'+""" class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 4).comment.text+'"'+""">"""+str(sheet1.cell(i, 4).value)+"""</td>
          <td class=xl3 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 5).value+'"'+""" class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 6).comment.text+'"'+""">"""+str(sheet1.cell(i, 6).value)+"""</td>
         </tr>
        """
    for i in range(277,313):
        message=message+""" 
         <tr height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl4 style='height:36.0pt;border-top:none'><output id="""+'"'+sheet1.cell(i, 1).value+'"'+""" height=48 class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 2).comment.text+'"'+""">"""+str(sheet1.cell(i, 2).value)+"""</td>
          <td class=xl4 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 3).value+'"'+""" class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 4).comment.text+'"'+""">"""+str(sheet1.cell(i, 4).value)+"""</td>
          <td class=xl4 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 5).value+'"'+""" class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 6).comment.text+'"'+""">"""+str(sheet1.cell(i, 6).value)+"""</td>
         </tr>
        """
    for i in range(313,331):
        message=message+""" 
         <tr height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl3 style='height:36.0pt;border-top:none'><output id="""+'"'+sheet1.cell(i, 1).value+'"'+""" height=48 class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 2).comment.text+'"'+""">"""+str(sheet1.cell(i, 2).value)+"""</td>
          <td class=xl3 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 3).value+'"'+""" class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 4).comment.text+'"'+""">"""+str(sheet1.cell(i, 4).value)+"""</td>
          <td class=xl3 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 5).value+'"'+""" class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 6).comment.text+'"'+""">"""+str(sheet1.cell(i, 6).value)+"""</td>
         </tr>
        """
    for i in range(331,337):
        message=message+""" 
         <tr height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl4 style='height:36.0pt;border-top:none'><output id="""+'"'+sheet1.cell(i, 1).value+'"'+""" height=48 class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 2).comment.text+'"'+""">"""+str(sheet1.cell(i, 2).value)+"""</td>
          <td class=xl4 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 3).value+'"'+""" class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 4).comment.text+'"'+""">"""+str(sheet1.cell(i, 4).value)+"""</td>
          <td class=xl4 style='border-top:none;border-left:none'><output id="""+'"'+sheet1.cell(i, 5).value+'"'+""" class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl4 style='border-top:none;border-left:none' title="""+'"'+sheet1.cell(i, 6).comment.text+'"'+""">"""+str(sheet1.cell(i, 6).value)+"""</td>
         </tr>
        """
    message=message+"""
     <![if supportMisalignedColumns]>
 <tr height=0 style='display:none'>
  <td width=53 style='width:40pt'></td>
  <td width=378 style='width:284pt'></td>
  <td width=53 style='width:40pt'></td>
  <td width=378 style='width:284pt'></td>
  <td width=53 style='width:40pt'></td>
  <td width=378 style='width:284pt'></td>
 </tr>
 <![endif]>
</table>
</div>

<!----------------------------->
<!--“从 EXCEL 发布网页”向导结束-->
<!----------------------------->
<script type="text/javascript">
function setCookie(cname,cvalue,exdays) {
  var d = new Date();
  d.setTime(d.getTime() + (exdays*24*60*60*1000));
  var expires = "expires=" + d.toGMTString();
  document.cookie = cname + "=" + cvalue + ";" + expires + ";path=/";
}

function getCookie(cname) {
  var name = cname + "=";
  var decodedCookie = decodeURIComponent(document.cookie);
  var ca = decodedCookie.split(';');
  for(var i = 0; i < ca.length; i++) {
    var c = ca[i];
    while (c.charAt(0) == ' ') {
      c = c.substring(1);
    }
    if (c.indexOf(name) == 0) {
      return c.substring(name.length, c.length);
    }
  }
  return "";
}

window.onload=function checkCookie() {
var out1="A";
var out2="B";
var out3="C";
var thiscorner="";
var origincode =["D","G","A","J","W","X","O","R","E","C","Q","M","K","I","Z","S","B","L","N","Y","H","F","T","P"];
var array =["a11","a13","a17","a19","a21","a23","a27","a29","a31","a33","a37","a39","a41","a43","a47","a49","a51","a53","a57","a59","a61","a63","a67","a69"];
var str=["JAD","JAE","JAF","JAG","JAH","JAI","JAW","JAM","JAN","JAO","JAP","JAQ","JAR","JAS","JAT","JAX","JAY","JAZ","JDA","JDB","JDC","JDG","JDH","JDI","JDW","JDM","JDN","JDO","JDP","JDQ","JDR","JDS","JDT","JDX","JDY","JDZ","JGA","JGB","JGC","JGD","JGE","JGF","JGW","JGM","JGN","JGO","JGP","JGQ","JGR","JGS","JGT","JGX","JGY","JGZ","JWA","JWB","JWC","JWD","JWE","JWF","JWG","JWH","JWI","JWO","JWP","JWQ","JWR","JWS","JWT","JWX","JWY","JWZ","JOA","JOB","JOC","JOD","JOE","JOF","JOG","JOH","JOI","JOW","JOM","JON","JOR","JOS","JOT","JOX","JOY","JOZ","JRA","JRB","JRC","JRD","JRE","JRF","JRG","JRH","JRI","JRW","JRM","JRN","JRO","JRP","JRQ","JRX","JRY","JRZ","JXA","JXB","JXC","JXD","JXE","JXF","JXG","JXH","JXI","JXW","JXM","JXN","JXO","JXP","JXQ","JXR","JXS","JXT","GAD","GAE","GAF","GAW","GAM","GAN","GAO","GAP","GAQ","GAR","GAS","GAT","GAX","GAY","GAZ","GDA","GDB","GDC","GDW","GDM","GDN","GDO","GDP","GDQ","GDR","GDS","GDT","GDX","GDY","GDZ","GWA","GWB","GWC","GWD","GWE","GWF","GWO","GWP","GWQ","GWR","GWS","GWT","GWX","GWY","GWZ","GOA","GOB","GOC","GOD","GOE","GOF","GOW","GOM","GON","GOR","GOS","GOT","GOX","GOY","GOZ","GRA","GRB","GRC","GRD","GRE","GRF","GRW","GRM","GRN","GRO","GRP","GRQ","GRX","GRY","GRZ","GXA","GXB","GXC","GXD","GXE","GXF","GXW","GXM","GXN","GXO","GXP","GXQ","GXR","GXS","GXT","ADW","ADM","ADN","ADO","ADP","ADQ","ADR","ADS","ADT","ADX","ADY","ADZ","AWD","AWE","AWF","AWO","AWP","AWQ","AWR","AWS","AWT","AWX","AWY","AWZ","AOD","AOE","AOF","AOW","AOM","AON","AOR","AOS","AOT","AOX","AOY","AOZ","ARD","ARE","ARF","ARW","ARM","ARN","ARO","ARP","ARQ","ARX","ARY","ARZ","AXD","AXE","AXF","AXW","AXM","AXN","AXO","AXP","AXQ","AXR","AXS","AXT","DWO","DWP","DWQ","DWR","DWS","DWT","DWX","DWY","DWZ","DOW","DOM","DON","DOR","DOS","DOT","DOX","DOY","DOZ","DRW","DRM","DRN","DRO","DRP","DRQ","DRX","DRY","DRZ","DXW","DXM","DXN","DXO","DXP","DXQ","DXR","DXS","DXT","XWO","XWP","XWQ","XWR","XWS","XWT","XOW","XOM","XON","XOR","XOS","XOT","XRW","XRM","XRN","XRO","XRP","XRQ","WOR","WOS","WOT","WRO","WRP","WRQ","JBD","JBE","JBF","JBG","JBH","JBI","JBW","JBM","JBN","JBO","JBP","JBQ","JBR","JBS","JBT","JBX","JBY","JBZ","JEA","JEB","JEC","JEG","JEH","JEI","JEW","JEM","JEN","JEO","JEP","JEQ","JER","JES","JET","JEX","JEY","JEZ","JHA","JHB","JHC","JHD","JHE","JHF","JHW","JHM","JHN","JHO","JHP","JHQ","JHR","JHS","JHT","JHX","JHY","JHZ","JMA","JMB","JMC","JMD","JME","JMF","JMG","JMH","JMI","JMO","JMP","JMQ","JMR","JMS","JMT","JMX","JMY","JMZ","JPA","JPB","JPC","JPD","JPE","JPF","JPG","JPH","JPI","JPW","JPM","JPN","JPR","JPS","JPT","JPX","JPY","JPZ","JSA","JSB","JSC","JSD","JSE","JSF","JSG","JSH","JSI","JSW","JSM","JSN","JSO","JSP","JSQ","JSX","JSY","JSZ","JYA","JYB","JYC","JYD","JYE","JYF","JYG","JYH","JYI","JYW","JYM","JYN","JYO","JYP","JYQ","JYR","JYS","JYT","GBD","GBE","GBF","GBW","GBM","GBN","GBO","GBP","GBQ","GBR","GBS","GBT","GBX","GBY","GBZ","GEA","GEB","GEC","GEW","GEM","GEN","GEO","GEP","GEQ","GER","GES","GET","GEX","GEY","GEZ","GMA","GMB","GMC","GMD","GME","GMF","GMO","GMP","GMQ","GMR","GMS","GMT","GMX","GMY","GMZ","GPA","GPB","GPC","GPD","GPE","GPF","GPW","GPM","GPN","GPR","GPS","GPT","GPX","GPY","GPZ","GSA","GSB","GSC","GSD","GSE","GSF","GSW","GSM","GSN","GSO","GSP","GSQ","GSX","GSY","GSZ","GYA","GYB","GYC","GYD","GYE","GYF","GYW","GYM","GYN","GYO","GYP","GYQ","GYR","GYS","GYT","AEW","AEM","AEN","AEO","AEP","AEQ","AER","AES","AET","AEX","AEY","AEZ","AMD","AME","AMF","AMO","AMP","AMQ","AMR","AMS","AMT","AMX","AMY","AMZ","APD","APE","APF","APW","APM","APN","APR","APS","APT","APX","APY","APZ","ASD","ASE","ASF","ASW","ASM","ASN","ASO","ASP","ASQ","ASX","ASY","ASZ","AYD","AYE","AYF","AYW","AYM","AYN","AYO","AYP","AYQ","AYR","AYS","AYT","DMO","DMP","DMQ","DMR","DMS","DMT","DMX","DMY","DMZ","DPW","DPM","DPN","DPR","DPS","DPT","DPX","DPY","DPZ","DSW","DSM","DSN","DSO","DSP","DSQ","DSX","DSY","DSZ","DYW","DYM","DYN","DYO","DYP","DYQ","DYR","DYS","DYT","XMO","XMP","XMQ","XMR","XMS","XMT","XPW","XPM","XPN","XPR","XPS","XPT","XSW","XSM","XSN","XSO","XSP","XSQ","WPR","WPS","WPT","WSO","WSP","WSQ","JCD","JCE","JCF","JCG","JCH","JCI","JCW","JCM","JCN","JCO","JCP","JCQ","JCR","JCS","JCT","JCX","JCY","JCZ","JFA","JFB","JFC","JFG","JFH","JFI","JFW","JFM","JFN","JFO","JFP","JFQ","JFR","JFS","JFT","JFX","JFY","JFZ","JIA","JIB","JIC","JID","JIE","JIF","JIW","JIM","JIN","JIO","JIP","JIQ","JIR","JIS","JIT","JIX","JIY","JIZ","JNA","JNB","JNC","JND","JNE","JNF","JNG","JNH","JNI","JNO","JNP","JNQ","JNR","JNS","JNT","JNX","JNY","JNZ","JQA","JQB","JQC","JQD","JQE","JQF","JQG","JQH","JQI","JQW","JQM","JQN","JQR","JQS","JQT","JQX","JQY","JQZ","JTA","JTB","JTC","JTD","JTE","JTF","JTG","JTH","JTI","JTW","JTM","JTN","JTO","JTP","JTQ","JTX","JTY","JTZ","JZA","JZB","JZC","JZD","JZE","JZF","JZG","JZH","JZI","JZW","JZM","JZN","JZO","JZP","JZQ","JZR","JZS","JZT","GCD","GCE","GCF","GCW","GCM","GCN","GCO","GCP","GCQ","GCR","GCS","GCT","GCX","GCY","GCZ","GFA","GFB","GFC","GFW","GFM","GFN","GFO","GFP","GFQ","GFR","GFS","GFT","GFX","GFY","GFZ","GNA","GNB","GNC","GND","GNE","GNF","GNO","GNP","GNQ","GNR","GNS","GNT","GNX","GNY","GNZ","GQA","GQB","GQC","GQD","GQE","GQF","GQW","GQM","GQN","GQR","GQS","GQT","GQX","GQY","GQZ","GTA","GTB","GTC","GTD","GTE","GTF","GTW","GTM","GTN","GTO","GTP","GTQ","GTX","GTY","GTZ","GZA","GZB","GZC","GZD","GZE","GZF","GZW","GZM","GZN","GZO","GZP","GZQ","GZR","GZS","GZT","AFW","AFM","AFN","AFO","AFP","AFQ","AFR","AFS","AFT","AFX","AFY","AFZ","AND","ANE","ANF","ANO","ANP","ANQ","ANR","ANS","ANT","ANX","ANY","ANZ","AQD","AQE","AQF","AQW","AQM","AQN","AQR","AQS","AQT","AQX","AQY","AQZ","ATD","ATE","ATF","ATW","ATM","ATN","ATO","ATP","ATQ","ATX","ATY","ATZ","AZD","AZE","AZF","AZW","AZM","AZN","AZO","AZP","AZQ","AZR","AZS","AZT","DNO","DNP","DNQ","DNR","DNS","DNT","DNX","DNY","DNZ","DQW","DQM","DQN","DQR","DQS","DQT","DQX","DQY","DQZ","DTW","DTM","DTN","DTO","DTP","DTQ","DTX","DTY","DTZ","DZW","DZM","DZN","DZO","DZP","DZQ","DZR","DZS","DZT","XNO","XNP","XNQ","XNR","XNS","XNT","XQW","XQM","XQN","XQR","XQS","XQT","XTW","XTM","XTN","XTO","XTP","XTQ","WQR","WQS","WQT","WTO","WTP","WTQ"];
var arrayout=[];
var len = array.length;
var strlen=str.length;
for (i=0 ; i < len; i++) { 
     if (getCookie(array[i]) != "") 
          arrayout[i]=getCookie(array[i]);
     else
          arrayout[i]=origincode[i];
}
for (stri=0 ; stri<strlen; stri++) { 
   res1 = str[stri].substring(0,1);
   res2 = str[stri].substring(1,2);
   res3 = str[stri].substring(2,3);
   for (i=0 ; i < len; i++) { 
        if (res1==origincode[i])
           out1=arrayout[i];
        if (res2==origincode[i])
           out2=arrayout[i];
        if (res3==origincode[i])
           out3=arrayout[i];
   }
   byid(str[stri]).value=out1.concat("",out2.concat("",out3));
}
}
 
function byid(id)
{
return document.getElementById(id);
}
<!--避免下面用document……麻烦可以简便调用-->
function getsum()
{
var array =["a11","a13","a17","a19","a21","a23","a27","a29","a31","a33","a37","a39","a41","a43","a47","a49","a51","a53","a57","a59","a61","a63","a67","a69"];
var len = array.length;
for (i=0 ; i < len; i++) { 
     setCookie(array[i], byid(array[i]).value, 30);
}
}	

function wu()
{
}	
</script>
</body>
</html>
    
    """
    f.write(message)
    f.close()


def make_each(sheetname):

    workbook = load_workbook(filename, data_only=True)
    if "1008" in filename:
        sheet1 = workbook[sheetname]
    GEN_HTML = sheetname+".html"
    f = open(GEN_HTML, 'w', encoding='utf-8')
    message = """
<html xmlns:v="urn:schemas-microsoft-com:vml"
xmlns:o="urn:schemas-microsoft-com:office:office"
xmlns:x="urn:schemas-microsoft-com:office:excel"
xmlns="http://www.w3.org/TR/REC-html40">

    <head>
        <script type="text/javascript">
//在页面未加载完毕之前显示的loading Html自定义内容
var _LoadingHtml = `<div id="loadingDiv">页面加载中，请等待...</div>`;
//呈现loading效果
document.write(_LoadingHtml);
//监听加载状态改变
document.body.style.display = "none";
document.onreadystatechange = completeLoading;
   
//加载状态为complete时移除loading效果
function completeLoading() {
    if (document.readyState == "complete") {
        document.body.style.display = "block";
        var loadingMask = document.getElementById('loadingDiv');
        loadingMask.parentNode.removeChild(loadingMask);
    } 
   else
       document.body.style.display = "none";
}
</script>

<title>【噩梦版】WYF's 全缓冲角1008</title>
<meta http-equiv=Content-Type content="text/html; charset=utf-8">
<meta name=ProgId content=Excel.Sheet>
<meta name=Generator content="Microsoft Excel 15">
<link rel=File-List href="UFR.files/filelist.xml">
<!--[if !mso]>
<style>
v\:* {behavior:url(#default#VML);}
o\:* {behavior:url(#default#VML);}
x\:* {behavior:url(#default#VML);}
.shape {behavior:url(#default#VML);}
</style>
<![endif]-->
<style id="UFR_Styles">
<!--table
	{mso-displayed-decimal-separator:"\.";
	mso-displayed-thousand-separator:"\,";}
.font1
	{color:black;
	font-size:13.3pt;
	font-weight:400;
	font-style:normal;
	text-decoration:none;
	font-family:Tahoma, sans-serif;
	mso-font-charset:0;}
.xl1
	{padding:0px;
	mso-ignore:padding;
	color:windowtext;
	font-size:13.3pt;
	font-weight:400;
	font-style:normal;
	text-decoration:none;
	font-family:宋体;
	mso-generic-font-family:auto;
	mso-font-charset:134;
	mso-number-format:General;
	text-align:general;
	vertical-align:middle;
	mso-background-source:auto;
	mso-pattern:auto;
	mso-protection:unlocked hidden;
	white-space:nowrap;}
.xl2
	{padding:0px;
	mso-ignore:padding;
	color:windowtext;
	font-size:13.3pt;
	font-weight:400;
	font-style:normal;
	text-decoration:none;
	font-family:宋体;
	mso-generic-font-family:auto;
	mso-font-charset:134;
	mso-number-format:General;
	text-align:general;
	vertical-align:middle;
	background:white;
	mso-pattern:black none;
	mso-protection:unlocked hidden;
	white-space:nowrap;}
.xl3
	{padding:0px;
	mso-ignore:padding;
	color:black;
	font-size:13.3pt;
	font-weight:400;
	font-style:normal;
	text-decoration:none;
	font-family:Tahoma, sans-serif;
	mso-font-charset:0;
	mso-number-format:General;
	text-align:center;
	vertical-align:middle;
	border:.5pt solid windowtext;
	background:white;
	mso-pattern:black none;
	mso-protection:locked hidden;
	white-space:nowrap;}
.xl4
	{padding:0px;
	mso-ignore:padding;
	color:black;
	font-size:13.3pt;
	font-weight:400;
	font-style:normal;
	text-decoration:none;
	font-family:Tahoma, sans-serif;
	mso-font-charset:0;
	mso-number-format:General;
	text-align:center;
	vertical-align:middle;
	border:.5pt solid windowtext;
	background:#D9D9D9;
	mso-pattern:black none;
	mso-protection:locked hidden;
	white-space:nowrap;}
ruby
	{ruby-align:left;}
rt
	{color:windowtext;
	font-size:9.0pt;
	font-weight:400;
	font-style:normal;
	text-decoration:none;
	font-family:宋体;
	mso-generic-font-family:auto;
	mso-font-charset:134;
	mso-char-type:none;}
-->
</style>
</head>

<body>
<!--[if !excel]>　　<![endif]-->
<!--下列信息由 Microsoft Excel 的发布为网页向导生成。-->
<!--如果同一条目从 Excel 中重新发布，则所有位于 DIV 标记之间的信息均将被替换。-->
<!----------------------------->
<!--“从 EXCEL 发布网页”向导开始-->
<!----------------------------->

<div id="UFR" align=center x:publishsource="Excel">

<table border=0 cellpadding=0 cellspacing=0 width=1293 class=xl2
 style='border-collapse:collapse;table-layout:fixed;width:923.4pt'>
 <col class=xl2 width=53 style='mso-width-source:userset;mso-width-alt:
 1685;width:38pt'>
 <col class=xl2 width=378 style='mso-width-source:userset;mso-width-alt:
 12096;width:269.8pt'>
 <col class=xl2 width=53 style='mso-width-source:userset;mso-width-alt:
 1685;width:38pt'>
 <col class=xl2 width=378 style='mso-width-source:userset;mso-width-alt:
 12096;width:269.8pt'>
 <col class=xl2 width=53 style='mso-width-source:userset;mso-width-alt:
 1685;width:38pt'>
 <col class=xl2 width=378 style='mso-width-source:userset;mso-width-alt:
 12096;width:269.8pt'>
    """
    for colori in range(0,3):
        for i in range(36*colori+1, 36*colori+19):
            message = message + """ 
              <tr class=xl1 height=48 style='mso-height-source:userset;height:36.0pt'>
              <td height=48 class=xl3 width=53 style='height:36.0pt;width:40pt'><output id=""" + '"' + sheet1.cell(i,1).value + '"' + """ height=48 class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
              <td class=xl3 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,2).comment.text + '"' + """>""" + str(sheet1.cell(i, 2).value) + """</td>
              <td class=xl3 width=53 style='border-left:none;width:40pt'><output id=""" + '"' + sheet1.cell(i, 3).value + '"' + """ class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
              <td class=xl3 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,4).comment.text + '"' + """>""" + str(sheet1.cell(i, 4).value) + """</td>
              <td class=xl3 width=53 style='border-left:none;width:40pt'><output id=""" + '"' + sheet1.cell(i, 5).value + '"' + """ class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
              <td class=xl3 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,6).comment.text + '"' + """>""" + str(sheet1.cell(i, 6).value) + """</td>
             </tr>
            """
        for i in range(36*colori+19, 36*colori+37):
            message = message + """ 
              <tr class=xl1 height=48 style='mso-height-source:userset;height:36.0pt'>
              <td height=48 class=xl4 width=53 style='height:36.0pt;width:40pt'><output id=""" + '"' + sheet1.cell(i,1).value + '"' + """ height=48 class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
              <td class=xl4 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,2).comment.text + '"' + """>""" + str(sheet1.cell(i, 2).value) + """</td>
              <td class=xl4 width=53 style='border-left:none;width:40pt'><output id=""" + '"' + sheet1.cell(i, 3).value + '"' + """ class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
              <td class=xl4 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,4).comment.text + '"' + """>""" + str(sheet1.cell(i, 4).value) + """</td>
              <td class=xl4 width=53 style='border-left:none;width:40pt'><output id=""" + '"' + sheet1.cell(i, 5).value + '"' + """ class=xl4 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
              <td class=xl4 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,6).comment.text + '"' + """>""" + str(sheet1.cell(i, 6).value) + """</td>
             </tr>
            """
    for i in range(36 * 3 + 1, 36 * 3 + 19):
        message = message + """ 
          <tr class=xl1 height=48 style='mso-height-source:userset;height:36.0pt'>
          <td height=48 class=xl3 width=53 style='height:36.0pt;width:40pt'><output id=""" + '"' + sheet1.cell(i,1).value + '"' + """ height=48 class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,2).comment.text + '"' + """>""" + str(sheet1.cell(i, 2).value) + """</td>
          <td class=xl3 width=53 style='border-left:none;width:40pt'><output id=""" + '"' + sheet1.cell(i,3).value + '"' + """ class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,4).comment.text + '"' + """>""" + str(sheet1.cell(i, 4).value) + """</td>
          <td class=xl3 width=53 style='border-left:none;width:40pt'><output id=""" + '"' + sheet1.cell(i,5).value + '"' + """ class=xl3 style='border-left:none;border-right:none;border-top:none;border-bottom:none'></td>
          <td class=xl3 width=378 style='border-left:none;width:284pt' title=""" + '"' + sheet1.cell(i,6).comment.text + '"' + """>""" + str(sheet1.cell(i, 6).value) + """</td>
         </tr>
        """
    message = message + """
 <![if supportMisalignedColumns]>
 <tr height=0 style='display:none'>
  <td width=53 style='width:40pt'></td>
  <td width=378 style='width:284pt'></td>
  <td width=53 style='width:40pt'></td>
  <td width=378 style='width:284pt'></td>
  <td width=53 style='width:40pt'></td>
  <td width=378 style='width:284pt'></td>
 </tr>
 <![endif]>
</table>

</div>


<!----------------------------->
<!--“从 EXCEL 发布网页”向导结束-->
<!----------------------------->
<script type="text/javascript">
function setCookie(cname,cvalue,exdays) {
  var d = new Date();
  d.setTime(d.getTime() + (exdays*24*60*60*1000));
  var expires = "expires=" + d.toGMTString();
  document.cookie = cname + "=" + cvalue + ";" + expires + ";path=/";
}

function getCookie(cname) {
  var name = cname + "=";
  var decodedCookie = decodeURIComponent(document.cookie);
  var ca = decodedCookie.split(';');
  for(var i = 0; i < ca.length; i++) {
    var c = ca[i];
    while (c.charAt(0) == ' ') {
      c = c.substring(1);
    }
    if (c.indexOf(name) == 0) {
      return c.substring(name.length, c.length);
    }
  }
  return "";
}

window.onload=function checkCookie() {
var out1="";
var out2="";
var out3="";
var thiscorner="";
var origincode =["D","G","A","J","W","X","O","R","E","C","Q","M","K","I","Z","S","B","L","N","Y","H","F","T","P"];
var array =["a11","a13","a17","a19","a21","a23","a27","a29","a31","a33","a37","a39","a41","a43","a47","a49","a51","a53","a57","a59","a61","a63","a67","a69"];
"""
    if sheetname=="UFR":
        message = message + """
var str=["AD","AE","AF","AG","AH","AI","AW","AM","AN","AO","AP","AQ","AR","AS","AT","AX","AY","AZ","DA","DB","DC","DG","DH","DI","DW","DM","DN","DO","DP","DQ","DR","DS","DT","DX","DY","DZ","GA","GB","GC","GD","GE","GF","GW","GM","GN","GO","GP","GQ","GR","GS","GT","GX","GY","GZ","WA","WB","WC","WD","WE","WF","WG","WH","WI","WO","WP","WQ","WR","WS","WT","WX","WY","WZ","OA","OB","OC","OD","OE","OF","OG","OH","OI","OW","OM","ON","OR","OS","OT","OX","OY","OZ","RA","RB","RC","RD","RE","RF","RG","RH","RI","RW","RM","RN","RO","RP","RQ","RX","RY","RZ","XA","XB","XC","XD","XE","XF","XG","XH","XI","XW","XM","XN","XO","XP","XQ","XR","XS","XT","BD","BE","BF","BG","BH","BI","BW","BM","BN","BO","BP","BQ","BR","BS","BT","BX","BY","BZ","EA","EB","EC","EG","EH","EI","EW","EM","EN","EO","EP","EQ","ER","ES","ET","EX","EY","EZ","HA","HB","HC","HD","HE","HF","HW","HM","HN","HO","HP","HQ","HR","HS","HT","HX","HY","HZ","MA","MB","MC","MD","ME","MF","MG","MH","MI","MO","MP","MQ","MR","MS","MT","MX","MY","MZ","PA","PB","PC","PD","PE","PF","PG","PH","PI","PW","PM","PN","PR","PS","PT","PX","PY","PZ","SA","SB","SC","SD","SE","SF","SG","SH","SI","SW","SM","SN","SO","SP","SQ","SX","SY","SZ","YA","YB","YC","YD","YE","YF","YG","YH","YI","YW","YM","YN","YO","YP","YQ","YR","YS","YT","CD","CE","CF","CG","CH","CI","CW","CM","CN","CO","CP","CQ","CR","CS","CT","CX","CY","CZ","FA","FB","FC","FG","FH","FI","FW","FM","FN","FO","FP","FQ","FR","FS","FT","FX","FY","FZ","IA","IB","IC","ID","IE","IF","IW","IM","IN","IO","IP","IQ","IR","IS","IT","IX","IY","IZ","NA","NB","NC","ND","NE","NF","NG","NH","NI","NO","NP","NQ","NR","NS","NT","NX","NY","NZ","QA","QB","QC","QD","QE","QF","QG","QH","QI","QW","QM","QN","QR","QS","QT","QX","QY","QZ","TA","TB","TC","TD","TE","TF","TG","TH","TI","TW","TM","TN","TO","TP","TQ","TX","TY","TZ","ZA","ZB","ZC","ZD","ZE","ZF","ZG","ZH","ZI","ZW","ZM","ZN","ZO","ZP","ZQ","ZR","ZS","ZT"]
"""
    if sheetname == "UBR":
        message = message + """
var str=["AD","AE","AF","AJ","AK","AL","AW","AM","AN","AO","AP","AQ","AR","AS","AT","AX","AY","AZ","DA","DB","DC","DJ","DK","DL","DW","DM","DN","DO","DP","DQ","DR","DS","DT","DX","DY","DZ","JA","JB","JC","JD","JE","JF","JW","JM","JN","JO","JP","JQ","JR","JS","JT","JX","JY","JZ","WA","WB","WC","WD","WE","WF","WJ","WK","WL","WO","WP","WQ","WR","WS","WT","WX","WY","WZ","OA","OB","OC","OD","OE","OF","OJ","OK","OL","OW","OM","ON","OR","OS","OT","OX","OY","OZ","RA","RB","RC","RD","RE","RF","RJ","RK","RL","RW","RM","RN","RO","RP","RQ","RX","RY","RZ","XA","XB","XC","XD","XE","XF","XJ","XK","XL","XW","XM","XN","XO","XP","XQ","XR","XS","XT","BD","BE","BF","BJ","BK","BL","BW","BM","BN","BO","BP","BQ","BR","BS","BT","BX","BY","BZ","EA","EB","EC","EJ","EK","EL","EW","EM","EN","EO","EP","EQ","ER","ES","ET","EX","EY","EZ","KA","KB","KC","KD","KE","KF","KW","KM","KN","KO","KP","KQ","KR","KS","KT","KX","KY","KZ","MA","MB","MC","MD","ME","MF","MJ","MK","ML","MO","MP","MQ","MR","MS","MT","MX","MY","MZ","PA","PB","PC","PD","PE","PF","PJ","PK","PL","PW","PM","PN","PR","PS","PT","PX","PY","PZ","SA","SB","SC","SD","SE","SF","SJ","SK","SL","SW","SM","SN","SO","SP","SQ","SX","SY","SZ","YA","YB","YC","YD","YE","YF","YJ","YK","YL","YW","YM","YN","YO","YP","YQ","YR","YS","YT","CD","CE","CF","CJ","CK","CL","CW","CM","CN","CO","CP","CQ","CR","CS","CT","CX","CY","CZ","FA","FB","FC","FJ","FK","FL","FW","FM","FN","FO","FP","FQ","FR","FS","FT","FX","FY","FZ","LA","LB","LC","LD","LE","LF","LW","LM","LN","LO","LP","LQ","LR","LS","LT","LX","LY","LZ","NA","NB","NC","ND","NE","NF","NJ","NK","NL","NO","NP","NQ","NR","NS","NT","NX","NY","NZ","QA","QB","QC","QD","QE","QF","QJ","QK","QL","QW","QM","QN","QR","QS","QT","QX","QY","QZ","TA","TB","TC","TD","TE","TF","TJ","TK","TL","TW","TM","TN","TO","TP","TQ","TX","TY","TZ","ZA","ZB","ZC","ZD","ZE","ZF","ZJ","ZK","ZL","ZW","ZM","ZN","ZO","ZP","ZQ","ZR","ZS","ZT"]
 """
    if sheetname == "UFL":
        message = message + """
var str=["DG","DH","DI","DJ","DK","DL","DW","DM","DN","DO","DP","DQ","DR","DS","DT","DX","DY","DZ","GD","GE","GF","GJ","GK","GL","GW","GM","GN","GO","GP","GQ","GR","GS","GT","GX","GY","GZ","JD","JE","JF","JG","JH","JI","JW","JM","JN","JO","JP","JQ","JR","JS","JT","JX","JY","JZ","WD","WE","WF","WG","WH","WI","WJ","WK","WL","WO","WP","WQ","WR","WS","WT","WX","WY","WZ","OD","OE","OF","OG","OH","OI","OJ","OK","OL","OW","OM","ON","OR","OS","OT","OX","OY","OZ","RD","RE","RF","RG","RH","RI","RJ","RK","RL","RW","RM","RN","RO","RP","RQ","RX","RY","RZ","XD","XE","XF","XG","XH","XI","XJ","XK","XL","XW","XM","XN","XO","XP","XQ","XR","XS","XT","EG","EH","EI","EJ","EK","EL","EW","EM","EN","EO","EP","EQ","ER","ES","ET","EX","EY","EZ","HD","HE","HF","HJ","HK","HL","HW","HM","HN","HO","HP","HQ","HR","HS","HT","HX","HY","HZ","KD","KE","KF","KG","KH","KI","KW","KM","KN","KO","KP","KQ","KR","KS","KT","KX","KY","KZ","MD","ME","MF","MG","MH","MI","MJ","MK","ML","MO","MP","MQ","MR","MS","MT","MX","MY","MZ","PD","PE","PF","PG","PH","PI","PJ","PK","PL","PW","PM","PN","PR","PS","PT","PX","PY","PZ","SD","SE","SF","SG","SH","SI","SJ","SK","SL","SW","SM","SN","SO","SP","SQ","SX","SY","SZ","YD","YE","YF","YG","YH","YI","YJ","YK","YL","YW","YM","YN","YO","YP","YQ","YR","YS","YT","FG","FH","FI","FJ","FK","FL","FW","FM","FN","FO","FP","FQ","FR","FS","FT","FX","FY","FZ","ID","IE","IF","IJ","IK","IL","IW","IM","IN","IO","IP","IQ","IR","IS","IT","IX","IY","IZ","LD","LE","LF","LG","LH","LI","LW","LM","LN","LO","LP","LQ","LR","LS","LT","LX","LY","LZ","ND","NE","NF","NG","NH","NI","NJ","NK","NL","NO","NP","NQ","NR","NS","NT","NX","NY","NZ","QD","QE","QF","QG","QH","QI","QJ","QK","QL","QW","QM","QN","QR","QS","QT","QX","QY","QZ","TD","TE","TF","TG","TH","TI","TJ","TK","TL","TW","TM","TN","TO","TP","TQ","TX","TY","TZ","ZD","ZE","ZF","ZG","ZH","ZI","ZJ","ZK","ZL","ZW","ZM","ZN","ZO","ZP","ZQ","ZR","ZS","ZT"]
"""
    if sheetname == "UBL":
        message = message + """
var str=["AG","AH","AI","AJ","AK","AL","AW","AM","AN","AO","AP","AQ","AR","AS","AT","AX","AY","AZ","GA","GB","GC","GJ","GK","GL","GW","GM","GN","GO","GP","GQ","GR","GS","GT","GX","GY","GZ","JA","JB","JC","JG","JH","JI","JW","JM","JN","JO","JP","JQ","JR","JS","JT","JX","JY","JZ","WA","WB","WC","WG","WH","WI","WJ","WK","WL","WO","WP","WQ","WR","WS","WT","WX","WY","WZ","OA","OB","OC","OG","OH","OI","OJ","OK","OL","OW","OM","ON","OR","OS","OT","OX","OY","OZ","RA","RB","RC","RG","RH","RI","RJ","RK","RL","RW","RM","RN","RO","RP","RQ","RX","RY","RZ","XA","XB","XC","XG","XH","XI","XJ","XK","XL","XW","XM","XN","XO","XP","XQ","XR","XS","XT","BG","BH","BI","BJ","BK","BL","BW","BM","BN","BO","BP","BQ","BR","BS","BT","BX","BY","BZ","HA","HB","HC","HJ","HK","HL","HW","HM","HN","HO","HP","HQ","HR","HS","HT","HX","HY","HZ","KA","KB","KC","KG","KH","KI","KW","KM","KN","KO","KP","KQ","KR","KS","KT","KX","KY","KZ","MA","MB","MC","MG","MH","MI","MJ","MK","ML","MO","MP","MQ","MR","MS","MT","MX","MY","MZ","PA","PB","PC","PG","PH","PI","PJ","PK","PL","PW","PM","PN","PR","PS","PT","PX","PY","PZ","SA","SB","SC","SG","SH","SI","SJ","SK","SL","SW","SM","SN","SO","SP","SQ","SX","SY","SZ","YA","YB","YC","YG","YH","YI","YJ","YK","YL","YW","YM","YN","YO","YP","YQ","YR","YS","YT","CG","CH","CI","CJ","CK","CL","CW","CM","CN","CO","CP","CQ","CR","CS","CT","CX","CY","CZ","IA","IB","IC","IJ","IK","IL","IW","IM","IN","IO","IP","IQ","IR","IS","IT","IX","IY","IZ","LA","LB","LC","LG","LH","LI","LW","LM","LN","LO","LP","LQ","LR","LS","LT","LX","LY","LZ","NA","NB","NC","NG","NH","NI","NJ","NK","NL","NO","NP","NQ","NR","NS","NT","NX","NY","NZ","QA","QB","QC","QG","QH","QI","QJ","QK","QL","QW","QM","QN","QR","QS","QT","QX","QY","QZ","TA","TB","TC","TG","TH","TI","TJ","TK","TL","TW","TM","TN","TO","TP","TQ","TX","TY","TZ","ZA","ZB","ZC","ZG","ZH","ZI","ZJ","ZK","ZL","ZW","ZM","ZN","ZO","ZP","ZQ","ZR","ZS","ZT"]
"""
    if sheetname == "DFR":
        message = message + """
var str=["AD","AE","AF","AG","AH","AI","AJ","AK","AL","AW","AM","AN","AO","AP","AQ","AR","AS","AT","DA","DB","DC","DG","DH","DI","DJ","DK","DL","DW","DM","DN","DO","DP","DQ","DR","DS","DT","GA","GB","GC","GD","GE","GF","GJ","GK","GL","GW","GM","GN","GO","GP","GQ","GR","GS","GT","JA","JB","JC","JD","JE","JF","JG","JH","JI","JW","JM","JN","JO","JP","JQ","JR","JS","JT","WA","WB","WC","WD","WE","WF","WG","WH","WI","WJ","WK","WL","WO","WP","WQ","WR","WS","WT","OA","OB","OC","OD","OE","OF","OG","OH","OI","OJ","OK","OL","OW","OM","ON","OR","OS","OT","RA","RB","RC","RD","RE","RF","RG","RH","RI","RJ","RK","RL","RW","RM","RN","RO","RP","RQ","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BW","BM","BN","BO","BP","BQ","BR","BS","BT","EA","EB","EC","EG","EH","EI","EJ","EK","EL","EW","EM","EN","EO","EP","EQ","ER","ES","ET","HA","HB","HC","HD","HE","HF","HJ","HK","HL","HW","HM","HN","HO","HP","HQ","HR","HS","HT","KA","KB","KC","KD","KE","KF","KG","KH","KI","KW","KM","KN","KO","KP","KQ","KR","KS","KT","MA","MB","MC","MD","ME","MF","MG","MH","MI","MJ","MK","ML","MO","MP","MQ","MR","MS","MT","PA","PB","PC","PD","PE","PF","PG","PH","PI","PJ","PK","PL","PW","PM","PN","PR","PS","PT","SA","SB","SC","SD","SE","SF","SG","SH","SI","SJ","SK","SL","SW","SM","SN","SO","SP","SQ","CD","CE","CF","CG","CH","CI","CJ","CK","CL","CW","CM","CN","CO","CP","CQ","CR","CS","CT","FA","FB","FC","FG","FH","FI","FJ","FK","FL","FW","FM","FN","FO","FP","FQ","FR","FS","FT","IA","IB","IC","ID","IE","IF","IJ","IK","IL","IW","IM","IN","IO","IP","IQ","IR","IS","IT","LA","LB","LC","LD","LE","LF","LG","LH","LI","LW","LM","LN","LO","LP","LQ","LR","LS","LT","NA","NB","NC","ND","NE","NF","NG","NH","NI","NJ","NK","NL","NO","NP","NQ","NR","NS","NT","QA","QB","QC","QD","QE","QF","QG","QH","QI","QJ","QK","QL","QW","QM","QN","QR","QS","QT","TA","TB","TC","TD","TE","TF","TG","TH","TI","TJ","TK","TL","TW","TM","TN","TO","TP","TQ"]
"""
    if sheetname == "DBR":
        message = message + """
var str=["AD","AE","AF","AG","AH","AI","AJ","AK","AL","AW","AM","AN","AO","AP","AQ","AX","AY","AZ","DA","DB","DC","DG","DH","DI","DJ","DK","DL","DW","DM","DN","DO","DP","DQ","DX","DY","DZ","GA","GB","GC","GD","GE","GF","GJ","GK","GL","GW","GM","GN","GO","GP","GQ","GX","GY","GZ","JA","JB","JC","JD","JE","JF","JG","JH","JI","JW","JM","JN","JO","JP","JQ","JX","JY","JZ","WA","WB","WC","WD","WE","WF","WG","WH","WI","WJ","WK","WL","WO","WP","WQ","WX","WY","WZ","OA","OB","OC","OD","OE","OF","OG","OH","OI","OJ","OK","OL","OW","OM","ON","OX","OY","OZ","XA","XB","XC","XD","XE","XF","XG","XH","XI","XJ","XK","XL","XW","XM","XN","XO","XP","XQ","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BW","BM","BN","BO","BP","BQ","BX","BY","BZ","EA","EB","EC","EG","EH","EI","EJ","EK","EL","EW","EM","EN","EO","EP","EQ","EX","EY","EZ","HA","HB","HC","HD","HE","HF","HJ","HK","HL","HW","HM","HN","HO","HP","HQ","HX","HY","HZ","KA","KB","KC","KD","KE","KF","KG","KH","KI","KW","KM","KN","KO","KP","KQ","KX","KY","KZ","MA","MB","MC","MD","ME","MF","MG","MH","MI","MJ","MK","ML","MO","MP","MQ","MX","MY","MZ","PA","PB","PC","PD","PE","PF","PG","PH","PI","PJ","PK","PL","PW","PM","PN","PX","PY","PZ","YA","YB","YC","YD","YE","YF","YG","YH","YI","YJ","YK","YL","YW","YM","YN","YO","YP","YQ","CD","CE","CF","CG","CH","CI","CJ","CK","CL","CW","CM","CN","CO","CP","CQ","CX","CY","CZ","FA","FB","FC","FG","FH","FI","FJ","FK","FL","FW","FM","FN","FO","FP","FQ","FX","FY","FZ","IA","IB","IC","ID","IE","IF","IJ","IK","IL","IW","IM","IN","IO","IP","IQ","IX","IY","IZ","LA","LB","LC","LD","LE","LF","LG","LH","LI","LW","LM","LN","LO","LP","LQ","LX","LY","LZ","NA","NB","NC","ND","NE","NF","NG","NH","NI","NJ","NK","NL","NO","NP","NQ","NX","NY","NZ","QA","QB","QC","QD","QE","QF","QG","QH","QI","QJ","QK","QL","QW","QM","QN","QX","QY","QZ","ZA","ZB","ZC","ZD","ZE","ZF","ZG","ZH","ZI","ZJ","ZK","ZL","ZW","ZM","ZN","ZO","ZP","ZQ"]
"""
    if sheetname == "DFL":
        message = message + """
var str=["AD","AE","AF","AG","AH","AI","AJ","AK","AL","AO","AP","AQ","AR","AS","AT","AX","AY","AZ","DA","DB","DC","DG","DH","DI","DJ","DK","DL","DO","DP","DQ","DR","DS","DT","DX","DY","DZ","GA","GB","GC","GD","GE","GF","GJ","GK","GL","GO","GP","GQ","GR","GS","GT","GX","GY","GZ","JA","JB","JC","JD","JE","JF","JG","JH","JI","JO","JP","JQ","JR","JS","JT","JX","JY","JZ","OA","OB","OC","OD","OE","OF","OG","OH","OI","OJ","OK","OL","OR","OS","OT","OX","OY","OZ","RA","RB","RC","RD","RE","RF","RG","RH","RI","RJ","RK","RL","RO","RP","RQ","RX","RY","RZ","XA","XB","XC","XD","XE","XF","XG","XH","XI","XJ","XK","XL","XO","XP","XQ","XR","XS","XT","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BO","BP","BQ","BR","BS","BT","BX","BY","BZ","EA","EB","EC","EG","EH","EI","EJ","EK","EL","EO","EP","EQ","ER","ES","ET","EX","EY","EZ","HA","HB","HC","HD","HE","HF","HJ","HK","HL","HO","HP","HQ","HR","HS","HT","HX","HY","HZ","KA","KB","KC","KD","KE","KF","KG","KH","KI","KO","KP","KQ","KR","KS","KT","KX","KY","KZ","PA","PB","PC","PD","PE","PF","PG","PH","PI","PJ","PK","PL","PR","PS","PT","PX","PY","PZ","SA","SB","SC","SD","SE","SF","SG","SH","SI","SJ","SK","SL","SO","SP","SQ","SX","SY","SZ","YA","YB","YC","YD","YE","YF","YG","YH","YI","YJ","YK","YL","YO","YP","YQ","YR","YS","YT","CD","CE","CF","CG","CH","CI","CJ","CK","CL","CO","CP","CQ","CR","CS","CT","CX","CY","CZ","FA","FB","FC","FG","FH","FI","FJ","FK","FL","FO","FP","FQ","FR","FS","FT","FX","FY","FZ","IA","IB","IC","ID","IE","IF","IJ","IK","IL","IO","IP","IQ","IR","IS","IT","IX","IY","IZ","LA","LB","LC","LD","LE","LF","LG","LH","LI","LO","LP","LQ","LR","LS","LT","LX","LY","LZ","QA","QB","QC","QD","QE","QF","QG","QH","QI","QJ","QK","QL","QR","QS","QT","QX","QY","QZ","TA","TB","TC","TD","TE","TF","TG","TH","TI","TJ","TK","TL","TO","TP","TQ","TX","TY","TZ","ZA","ZB","ZC","ZD","ZE","ZF","ZG","ZH","ZI","ZJ","ZK","ZL","ZO","ZP","ZQ","ZR","ZS","ZT"]
"""
    if sheetname == "DBL":
        message = message + """
var str=["AD","AE","AF","AG","AH","AI","AJ","AK","AL","AW","AM","AN","AR","AS","AT","AX","AY","AZ","DA","DB","DC","DG","DH","DI","DJ","DK","DL","DW","DM","DN","DR","DS","DT","DX","DY","DZ","GA","GB","GC","GD","GE","GF","GJ","GK","GL","GW","GM","GN","GR","GS","GT","GX","GY","GZ","JA","JB","JC","JD","JE","JF","JG","JH","JI","JW","JM","JN","JR","JS","JT","JX","JY","JZ","WA","WB","WC","WD","WE","WF","WG","WH","WI","WJ","WK","WL","WR","WS","WT","WX","WY","WZ","RA","RB","RC","RD","RE","RF","RG","RH","RI","RJ","RK","RL","RW","RM","RN","RX","RY","RZ","XA","XB","XC","XD","XE","XF","XG","XH","XI","XJ","XK","XL","XW","XM","XN","XR","XS","XT","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BW","BM","BN","BR","BS","BT","BX","BY","BZ","EA","EB","EC","EG","EH","EI","EJ","EK","EL","EW","EM","EN","ER","ES","ET","EX","EY","EZ","HA","HB","HC","HD","HE","HF","HJ","HK","HL","HW","HM","HN","HR","HS","HT","HX","HY","HZ","KA","KB","KC","KD","KE","KF","KG","KH","KI","KW","KM","KN","KR","KS","KT","KX","KY","KZ","MA","MB","MC","MD","ME","MF","MG","MH","MI","MJ","MK","ML","MR","MS","MT","MX","MY","MZ","SA","SB","SC","SD","SE","SF","SG","SH","SI","SJ","SK","SL","SW","SM","SN","SX","SY","SZ","YA","YB","YC","YD","YE","YF","YG","YH","YI","YJ","YK","YL","YW","YM","YN","YR","YS","YT","CD","CE","CF","CG","CH","CI","CJ","CK","CL","CW","CM","CN","CR","CS","CT","CX","CY","CZ","FA","FB","FC","FG","FH","FI","FJ","FK","FL","FW","FM","FN","FR","FS","FT","FX","FY","FZ","IA","IB","IC","ID","IE","IF","IJ","IK","IL","IW","IM","IN","IR","IS","IT","IX","IY","IZ","LA","LB","LC","LD","LE","LF","LG","LH","LI","LW","LM","LN","LR","LS","LT","LX","LY","LZ","NA","NB","NC","ND","NE","NF","NG","NH","NI","NJ","NK","NL","NR","NS","NT","NX","NY","NZ","TA","TB","TC","TD","TE","TF","TG","TH","TI","TJ","TK","TL","TW","TM","TN","TX","TY","TZ","ZA","ZB","ZC","ZD","ZE","ZF","ZG","ZH","ZI","ZJ","ZK","ZL","ZW","ZM","ZN","ZR","ZS","ZT"]
"""
    message=message+"""
var arrayout=[];
var len = array.length;
var strlen=str.length;
for (i=0 ; i < len; i++) { 
     if (getCookie(array[i]) != "") 
          arrayout[i]=getCookie(array[i]);
     else
          arrayout[i]=origincode[i];
}
for (stri=0 ; stri<strlen; stri++) { 
   res1 = str[stri].substring(0,1);
   res2 = str[stri].substring(1,2);
   for (i=0 ; i < len; i++) { 
        if (res1==origincode[i])
           out1=arrayout[i];
        if (res2==origincode[i])
           out2=arrayout[i];
   }
   byid(str[stri]).value=out1.concat("",out2);
}
}

function byid(id)
{
return document.getElementById(id);
}
<!--避免下面用document……麻烦可以简便调用-->
function getsum()
{
var array =["a11","a13","a17","a19","a21","a23","a27","a29","a31","a33","a37","a39","a41","a43","a47","a49","a51","a53","a57","a59","a61","a63","a67","a69"];
var len = array.length;
for (i=0 ; i < len; i++) { 
     setCookie(array[i], byid(array[i]).value, 30);
}
}	

function wu()
{
}	
</script>
</body>
</html>

    """
    f.write(message)
    f.close()

if __name__ == '__main__':
    make_each("UFR")
    make_each("UBR")
    make_each("UFL")
    make_each("UBL")
    make_each("DFR")
    make_each("DBR")
    make_each("DFL")
    make_each("DBL")
    toexit = input("Please press any key to continue")